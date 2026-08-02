"""Report generation: summaries, Azure baseline, and workbook formatting."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .mapping import azure_service_mapping


def dataframe(rows: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> pd.DataFrame:
    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=columns or [])


def create_asset_summary(assets_df: pd.DataFrame) -> pd.DataFrame:
    if assets_df.empty:
        return pd.DataFrame(columns=["Asset Type", "Count", "Azure Candidate", "Migration Note"])
    summary = assets_df.groupby("Asset Type").size().reset_index(name="Count")
    mappings = summary["Asset Type"].apply(azure_service_mapping)
    summary["Azure Candidate"] = mappings.apply(lambda value: value[0])
    summary["Migration Note"] = mappings.apply(lambda value: value[1])
    return summary.sort_values(["Count", "Asset Type"], ascending=[False, True]).reset_index(drop=True)


def create_azure_baseline(vm_df: pd.DataFrame, disk_df: pd.DataFrame, cluster_df: pd.DataFrame,
                          nodepool_df: pd.DataFrame, mig_df: pd.DataFrame, assets_df: pd.DataFrame,
                          sql_df: pd.DataFrame, storage_df: pd.DataFrame, bigquery_df: pd.DataFrame,
                          logging_df: pd.DataFrame, backupdr_df: pd.DataFrame, netapp_df: pd.DataFrame,
                          vmmanager_df: pd.DataFrame) -> pd.DataFrame:
    def numeric_sum(df: pd.DataFrame, column: str) -> float:
        if df.empty or column not in df.columns:
            return 0
        return float(pd.to_numeric(df[column], errors="coerce").fillna(0).sum())

    running = vm_df[vm_df["Status"] == "RUNNING"] if not vm_df.empty and "Status" in vm_df.columns else pd.DataFrame()
    stopped = vm_df[vm_df["Status"].isin(["TERMINATED", "STOPPED", "SUSPENDED"])] if not vm_df.empty and "Status" in vm_df.columns else pd.DataFrame()
    gke_nodes = vm_df[vm_df["Workload Association"] == "GKE Node"] if not vm_df.empty and "Workload Association" in vm_df.columns else pd.DataFrame()
    mig_vms = vm_df[vm_df["Managed Instance Group"].astype(str) != ""] if not vm_df.empty and "Managed Instance Group" in vm_df.columns else pd.DataFrame()

    sql_primaries = sql_df[sql_df["Instance Role"] == "CLOUD_SQL_INSTANCE"] if not sql_df.empty and "Instance Role" in sql_df.columns else pd.DataFrame()
    sql_replicas = sql_df[sql_df["Instance Role"] == "READ_REPLICA_INSTANCE"] if not sql_df.empty and "Instance Role" in sql_df.columns else pd.DataFrame()
    sql_ha = sql_df[sql_df["High Availability"] == "Yes"] if not sql_df.empty and "High Availability" in sql_df.columns else pd.DataFrame()

    netapp_volumes = netapp_df[netapp_df["Resource Type"] == "Volume"] if not netapp_df.empty and "Resource Type" in netapp_df.columns else pd.DataFrame()
    vm_patch = vmmanager_df[vmmanager_df["Resource Type"] == "Patch Deployment"] if not vmmanager_df.empty and "Resource Type" in vmmanager_df.columns else pd.DataFrame()

    metrics = [
        ("Projects", int(vm_df["Project"].nunique()) if not vm_df.empty and "Project" in vm_df.columns else 0, "Projects with discovered Compute Engine VMs"),
        ("Compute VMs - all", len(vm_df), "Includes running and stopped instances"),
        ("Compute VMs - running", len(running), "Configuration baseline for currently running instances"),
        ("Compute VMs - stopped/suspended", len(stopped), "Review whether these should migrate"),
        ("vCPU - all VMs", numeric_sum(vm_df, "vCPUs"), "Raw source allocation; not a utilization-based target"),
        ("Memory GiB - all VMs", numeric_sum(vm_df, "Memory GiB"), "Raw source allocation; not a utilization-based target"),
        ("vCPU - running VMs", numeric_sum(running, "vCPUs"), "Raw source allocation for running VMs"),
        ("Memory GiB - running VMs", numeric_sum(running, "Memory GiB"), "Raw source allocation for running VMs"),
        ("Persistent disks", len(disk_df), "Includes attached and unattached persistent disks"),
        ("Persistent disk capacity GiB", numeric_sum(disk_df, "Size GiB"), "Capacity only; collect IOPS/throughput/latency before tier selection"),
        ("External VM IP addresses", int(numeric_sum(vm_df, "External IP Count")), "Review public exposure and Azure ingress/egress design"),
        ("GKE clusters", len(cluster_df), "Candidate AKS clusters; consolidation requires architecture review"),
        ("GKE node pools", len(nodepool_df), "Candidate AKS node pools"),
        ("Discovered GKE nodes", len(gke_nodes), "Mapped through GKE instance groups and VM labels"),
        ("Managed instance groups", len(mig_df), "Candidate VM Scale Sets; GKE-owned groups normally become AKS node pools"),
        ("VMs in managed instance groups", len(mig_vms), "Includes GKE-backed managed groups"),
        ("Cloud SQL instances", len(sql_df), "MySQL, PostgreSQL, and SQL Server database instances"),
        ("Cloud SQL primary instances", len(sql_primaries), "Primary/standalone instances (excludes read replicas)"),
        ("Cloud SQL read replicas", len(sql_replicas), "Map to Azure read replicas where supported"),
        ("Cloud SQL HA instances", len(sql_ha), "Regional (HA) instances; map to Azure zone-redundant HA"),
        ("Cloud SQL storage GiB", numeric_sum(sql_df, "Data Disk Size GiB"), "Provisioned storage; collect growth and IOPS before tier selection"),
        ("Cloud Storage buckets", len(storage_df), "Object storage; collect stored bytes, object count, and egress via Cloud Monitoring"),
        ("BigQuery datasets", len(bigquery_df), "Analytics datasets; confirm query/slot patterns, pipelines, and governance"),
        ("BigQuery logical GiB (sampled)", numeric_sum(bigquery_df, "Total Logical GiB"), "Aggregated table logical size where computed; verify Size Complete column"),
        ("Cloud Logging resources", len(logging_df), "Log routing sinks and log storage buckets"),
        ("Backup and DR resources", len(backupdr_df), "Management servers, backup vaults, and backup plans"),
        ("NetApp volumes", len(netapp_volumes), "Map protocol, service level, and capacity to Azure NetApp Files"),
        ("NetApp volume capacity GiB", numeric_sum(netapp_volumes, "Capacity GiB"), "Provisioned volume capacity; confirm used vs provisioned and snapshots"),
        ("VM Manager patch deployments", len(vm_patch), "Map to Azure Update Manager schedules (Azure Arc for hybrid)"),
        ("Cloud Asset Inventory resources", len(assets_df), "Service count only; service-specific sizing is still required"),
    ]
    return pd.DataFrame(metrics, columns=["Metric", "Value", "Sizing / Migration Note"])


def create_recommendations() -> pd.DataFrame:
    rows = [
        (1, "Performance-based right-sizing", "Collect 14-30 days of CPU utilization, memory working set, disk IOPS/throughput/latency, and network throughput at suitable percentiles. Current recommendations use provisioned configuration only."),
        (2, "GKE workload inventory", "Export namespaces, deployments/statefulsets/daemonsets, pod requests and limits, HPA/VPA, PDBs, ingress, services, storage classes/PVCs, secrets/configmaps, CRDs, and add-ons. Node counts alone cannot size AKS accurately."),
        (3, "Database detail", "Cloud SQL engines, versions, editions, tiers (vCPU/RAM), storage, HA, replicas, backups, flags, and per-instance databases are now collected. Still add Spanner/Bigtable/Firestore/AlloyDB detail plus storage growth, IOPS, RPO/RTO, extensions, connections, and query performance."),
        (4, "Storage consumption", "Cloud Storage bucket configuration (location, class, versioning, lifecycle, retention, encryption, public access) is now collected. Still add stored bytes, object count, request rates, and egress via Cloud Monitoring before selecting Azure tier and redundancy."),
        (5, "Application dependencies", "Use application maps, VPC Flow Logs, load-balancer backends, DNS, service discovery, Kubernetes manifests, and owner interviews to identify communication paths and migration waves."),
        (6, "Network architecture", "Inventory CIDRs, routes, peering, Shared VPC, Cloud NAT, VPN/Interconnect, firewall rules, DNS, load balancers, certificates, bandwidth, latency, and overlapping address space."),
        (7, "Identity and security", "Map IAM roles, service-account usage, workload identity, KMS keys, secrets, org policies, Security Command Center findings, logging, retention, and regulatory controls to Entra ID and Azure Policy."),
        (8, "Availability and resilience", "Capture zones/regions, SLOs, maintenance windows, backup/restore tests, failover, RPO/RTO, autoscaling, quotas, and disaster-recovery topology."),
        (9, "Cost and licensing", "Add current GCP billing export, committed-use discounts, sustained-use effects, Windows/SQL/RHEL/SUSE licensing, support, backup, monitoring, data transfer, and Azure reservations/savings plans."),
        (10, "Azure SKU and quota validation", "Query the intended Azure region for VM/disk/AKS SKU availability, zone support, limits, quotas, and pricing before finalizing the BOQ."),
        (11, "PaaS-specific configuration", "Dedicated collectors now cover Cloud Storage, BigQuery, Cloud Logging, Backup and DR, NetApp Volumes, and VM Manager. Still add service-specific collectors for Cloud Run, Functions, Pub/Sub, Cloud Monitoring, Dataflow, Dataproc, Redis, DNS, KMS, Artifact Registry, and load balancers; asset counts are not enough for sizing."),
        (12, "Regional MIG completeness", "This script queries regional MIGs only in regions where VMs were discovered. Add an all-region scan if zero-instance regional MIGs must also be reported."),
    ]
    return pd.DataFrame(rows, columns=["Priority", "Feature to Add", "Why It Matters"])


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells[:2000]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 80))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    workbook.save(path)
