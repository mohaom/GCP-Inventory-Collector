#!/usr/bin/env python3
"""
GCP to Azure Migration Inventory & Baseline BOQ Generator
=========================================================

Discovers projects recursively under a GCP Organization, Folder, or Project;
inventories Cloud Asset Inventory resources; enriches Compute Engine VMs,
disks, NICs, Managed Instance Groups, GKE clusters/node pools, and Cloud SQL
database instances (MySQL, PostgreSQL, SQL Server); and dedicated collectors for
Cloud Storage, BigQuery, Cloud Logging, Backup and DR Service, NetApp Volumes,
and VM Manager (OS Config); and creates an Azure-oriented migration baseline.

Outputs (PREFIX defaults to "gcp_azure_inventory"):
  - PREFIX.xlsx
  - PREFIX_assets.csv
  - PREFIX_asset_summary.csv
  - PREFIX_compute.csv
  - PREFIX_disks.csv
  - PREFIX_gke_nodepools.csv
  - PREFIX_cloud_sql.csv
  - PREFIX_cloud_storage.csv
  - PREFIX_bigquery.csv
  - PREFIX_cloud_logging.csv
  - PREFIX_backup_dr.csv
  - PREFIX_netapp_volumes.csv
  - PREFIX_vm_manager.csv

Prerequisites:
  pip install google-cloud-asset google-api-python-client google-auth pandas openpyxl

Authentication:
  gcloud auth application-default login

Recommended APIs:
  cloudasset.googleapis.com
  cloudresourcemanager.googleapis.com
  compute.googleapis.com
  container.googleapis.com
  sqladmin.googleapis.com
  storage.googleapis.com
  bigquery.googleapis.com
  logging.googleapis.com
  backupdr.googleapis.com
  netapp.googleapis.com
  osconfig.googleapis.com

Examples:
  python gcp_azure_inventory.py --scope organizations/123456789
  python gcp_azure_inventory.py --scope folders/12345
  python gcp_azure_inventory.py --scope projects/my-project
  python gcp_azure_inventory.py --scope organizations/123456789 --output-prefix customer-prod

Important:
  Azure VM recommendations are configuration-based family suggestions, not final
  right-sizing. Final sizing should use 14-30 days of CPU, memory, disk IOPS,
  throughput, latency, and network utilization data plus Azure-region SKU checks.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict, deque
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from google.auth import default
from google.cloud import asset_v1
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CLOUD_PLATFORM_SCOPE = "https://www.googleapis.com/auth/cloud-platform"


# -----------------------------------------------------------------------------
# General helpers
# -----------------------------------------------------------------------------

def execute(request):
    """Execute a Google API request with a consistent user agent/error surface."""
    return request.execute(num_retries=3)


def basename(value: str) -> str:
    return value.rstrip("/").split("/")[-1] if value else ""


def extract_path_value(url: str, segment: str) -> str:
    if not url:
        return ""
    parts = url.rstrip("/").split("/")
    try:
        return parts[parts.index(segment) + 1]
    except (ValueError, IndexError):
        return ""


def zone_to_region(zone: str) -> str:
    return zone.rsplit("-", 1)[0] if zone and "-" in zone else ""


def compact_json(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, separators=(",", ":"), sort_keys=True, default=str)


def join_values(values: Iterable[Any], separator: str = ";") -> str:
    return separator.join(str(v) for v in values if v not in (None, ""))


def key_value_string(values: Dict[str, Any]) -> str:
    return ";".join(f"{k}={v}" for k, v in sorted(values.items()))


def metadata_to_dict(instance: Dict[str, Any]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for item in instance.get("metadata", {}).get("items", []) or []:
        key = item.get("key")
        if key:
            result[key] = item.get("value", "")
    return result


def infer_os(source_image: str, licenses: Iterable[str], metadata: Dict[str, str]) -> str:
    text = " ".join([source_image or "", *list(licenses), *metadata.values()]).lower()
    checks = [
        ("windows", "Windows Server"),
        ("sql-server", "Windows Server / SQL Server"),
        ("ubuntu-pro", "Ubuntu Pro"),
        ("ubuntu", "Ubuntu"),
        ("debian", "Debian"),
        ("rhel", "Red Hat Enterprise Linux"),
        ("red-hat", "Red Hat Enterprise Linux"),
        ("rocky", "Rocky Linux"),
        ("centos", "CentOS"),
        ("sles", "SUSE Linux Enterprise"),
        ("suse", "SUSE Linux Enterprise"),
        ("cos-cloud", "Container-Optimized OS"),
        ("container-optimized", "Container-Optimized OS"),
        ("fedora", "Fedora"),
        ("oracle-linux", "Oracle Linux"),
    ]
    for token, name in checks:
        if token in text:
            return name
    return "Unknown"


def parse_project_number(ancestors: Iterable[str]) -> str:
    for ancestor in ancestors:
        if ancestor.startswith("projects/"):
            return ancestor.split("/", 1)[1]
    return ""


def parse_custom_machine_type(machine_type: str) -> Tuple[Optional[int], Optional[float]]:
    # Example: custom-4-8192 or custom-4-8192-ext (memory is MiB).
    match = re.match(r"custom-(\d+)-(\d+)", machine_type or "")
    if not match:
        return None, None
    return int(match.group(1)), round(int(match.group(2)) / 1024.0, 2)


# -----------------------------------------------------------------------------
# Project and asset discovery
# -----------------------------------------------------------------------------

def discover_projects(scope: str, crm, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    """Recursively discover active projects below an org/folder, or return one project."""
    if scope.startswith("projects/"):
        project_ref = scope.split("/", 1)[1]
        try:
            # projects.get expects projects/PROJECT_NUMBER, but many environments also
            # resolve a project ID. If it does not, retain a minimal record.
            project = execute(crm.projects().get(name=scope))
            return [{
                "Project ID": project.get("projectId", project_ref),
                "Project Number": basename(project.get("name", "")),
                "Display Name": project.get("displayName", ""),
                "Parent": project.get("parent", ""),
                "State": project.get("state", "ACTIVE"),
                "Labels": key_value_string(project.get("labels", {})),
            }]
        except Exception as ex:
            errors.append({"Area": "Project discovery", "Project": project_ref, "Error": str(ex)})
            return [{
                "Project ID": project_ref,
                "Project Number": "",
                "Display Name": "",
                "Parent": "",
                "State": "UNKNOWN",
                "Labels": "",
            }]

    if not (scope.startswith("organizations/") or scope.startswith("folders/")):
        raise ValueError("--scope must be organizations/ID, folders/ID, or projects/ID")

    projects: List[Dict[str, Any]] = []
    queue: deque[str] = deque([scope])
    visited: set[str] = set()

    while queue:
        parent = queue.popleft()
        if parent in visited:
            continue
        visited.add(parent)

        # Direct child projects.
        token = None
        while True:
            try:
                kwargs = {"parent": parent, "pageSize": 1000, "showDeleted": False}
                if token:
                    kwargs["pageToken"] = token
                response = execute(crm.projects().list(**kwargs))
                for project in response.get("projects", []):
                    if project.get("state") == "DELETE_REQUESTED":
                        continue
                    projects.append({
                        "Project ID": project.get("projectId", ""),
                        "Project Number": basename(project.get("name", "")),
                        "Display Name": project.get("displayName", ""),
                        "Parent": project.get("parent", parent),
                        "State": project.get("state", ""),
                        "Labels": key_value_string(project.get("labels", {})),
                    })
                token = response.get("nextPageToken")
                if not token:
                    break
            except Exception as ex:
                errors.append({"Area": "Project discovery", "Project": parent, "Error": str(ex)})
                break

        # Child folders for recursion.
        token = None
        while True:
            try:
                kwargs = {"parent": parent, "pageSize": 1000, "showDeleted": False}
                if token:
                    kwargs["pageToken"] = token
                response = execute(crm.folders().list(**kwargs))
                for folder in response.get("folders", []):
                    if folder.get("state") != "DELETE_REQUESTED" and folder.get("name"):
                        queue.append(folder["name"])
                token = response.get("nextPageToken")
                if not token:
                    break
            except Exception as ex:
                errors.append({"Area": "Folder discovery", "Project": parent, "Error": str(ex)})
                break

    # De-duplicate by project ID while preserving the most complete record.
    unique: Dict[str, Dict[str, Any]] = {}
    for project in projects:
        pid = project.get("Project ID")
        if pid:
            unique[pid] = project
    return sorted(unique.values(), key=lambda row: row.get("Project ID", ""))


def list_assets(scope: str, project_number_to_id: Dict[str, str]) -> List[Dict[str, Any]]:
    client = asset_v1.AssetServiceClient()
    request = asset_v1.ListAssetsRequest(
        parent=scope,
        content_type=asset_v1.ContentType.RESOURCE,
        page_size=1000,
    )

    rows: List[Dict[str, Any]] = []
    for asset in client.list_assets(request=request):
        ancestors = list(asset.ancestors)
        project_number = parse_project_number(ancestors)
        resource = asset.resource
        rows.append({
            "Project ID": project_number_to_id.get(project_number, ""),
            "Project Number": project_number,
            "Asset Type": asset.asset_type,
            "Name": asset.name,
            "Display Name": getattr(resource, "discovery_name", ""),
            "Location": getattr(resource, "location", ""),
            "Version": getattr(resource, "version", ""),
            "Resource URL": getattr(resource, "discovery_document_uri", ""),
            "Ancestors": ";".join(ancestors),
        })
    return rows


# -----------------------------------------------------------------------------
# Azure mapping and sizing helpers
# -----------------------------------------------------------------------------

def azure_service_mapping(asset_type: str) -> Tuple[str, str]:
    mapping = [
        ("compute.googleapis.com/Instance", "Azure Virtual Machines / VM Scale Sets / AKS nodes", "Determine target from GKE and MIG association."),
        ("compute.googleapis.com/Disk", "Azure Managed Disks", "Map disk tier using capacity, IOPS, throughput, and latency."),
        ("compute.googleapis.com/Image", "Azure Compute Gallery", "Validate OS support and image licensing."),
        ("compute.googleapis.com/Snapshot", "Azure Disk Snapshots / Azure Backup", "Review retention and recovery requirements."),
        ("compute.googleapis.com/Network", "Azure Virtual Network", "Recreate CIDR, peering, DNS, routing, and segmentation."),
        ("compute.googleapis.com/Subnetwork", "Azure VNet Subnet", "Validate address-space overlap and delegated subnet needs."),
        ("compute.googleapis.com/Firewall", "Network Security Groups / Azure Firewall", "Translate allow/deny rules and service tags."),
        ("compute.googleapis.com/ForwardingRule", "Azure Load Balancer / Application Gateway / Front Door", "Target depends on L4/L7, internal/external, and global/regional scope."),
        ("compute.googleapis.com/BackendService", "Azure Load Balancer / Application Gateway backend pool", "Capture health probes, affinity, protocol, timeout, and capacity."),
        ("compute.googleapis.com/Address", "Azure Public IP", "Check regional/global and static/dynamic requirements."),
        ("compute.googleapis.com/Router", "Azure VPN Gateway / Route Server / NAT Gateway", "Review BGP, Cloud NAT, routes, and HA."),
        ("compute.googleapis.com/VpnTunnel", "Azure VPN Gateway connection", "Capture IKE/IPsec policy, bandwidth, and redundancy."),
        ("container.googleapis.com/Cluster", "Azure Kubernetes Service (AKS)", "Collect workloads, requests/limits, ingress, storage classes, and add-ons."),
        ("sqladmin.googleapis.com/Instance", "Azure SQL / Azure Database for PostgreSQL or MySQL", "Requires engine, version, vCPU, RAM, storage, IOPS, HA, replicas, and extensions."),
        ("storage.googleapis.com/Bucket", "Azure Blob Storage / ADLS Gen2", "Collect stored bytes, object count, access tier, versioning, lifecycle, and egress."),
        ("pubsub.googleapis.com/Topic", "Azure Service Bus / Event Grid / Event Hubs", "Choose by ordering, throughput, retention, and delivery semantics."),
        ("pubsub.googleapis.com/Subscription", "Azure Service Bus subscription / Event Grid subscription", "Capture filters, retry, dead-letter, ack deadline, and retention."),
        ("run.googleapis.com/Service", "Azure Container Apps / App Service", "Collect CPU/memory limits, concurrency, min/max instances, and networking."),
        ("cloudfunctions.googleapis.com/Function", "Azure Functions", "Collect runtime, trigger, memory, timeout, concurrency, and dependencies."),
        ("artifactregistry.googleapis.com/Repository", "Azure Container Registry", "Collect repository format, size, retention, and geo-replication."),
        ("bigquery.googleapis.com/Dataset", "Microsoft Fabric / Azure Synapse / ADLS Gen2", "Requires table sizes, query patterns, slots, pipelines, and governance."),
        ("bigquery.googleapis.com/Table", "Microsoft Fabric / Azure Synapse / ADLS Gen2", "Collect bytes, partitions, clustering, update frequency, and dependencies."),
        ("redis.googleapis.com/Instance", "Azure Managed Redis", "Collect tier, memory, throughput, clustering, persistence, and HA."),
        ("secretmanager.googleapis.com/Secret", "Azure Key Vault", "Map access policies/RBAC, rotation, versions, and private endpoints."),
        ("cloudkms.googleapis.com/CryptoKey", "Azure Key Vault / Managed HSM", "Map key type, protection level, rotation, and IAM."),
        ("dns.googleapis.com/ManagedZone", "Azure DNS / Private DNS", "Capture records, forwarding, private visibility, and DNSSEC."),
        ("iam.googleapis.com/ServiceAccount", "Microsoft Entra workload identity / Managed Identity", "Map identity usage and least-privilege roles."),
        ("logging.googleapis.com/LogSink", "Azure Monitor / Log Analytics / Event Hubs", "Map filters, destinations, retention, and volume."),
        ("monitoring.googleapis.com/AlertPolicy", "Azure Monitor alerts", "Translate signals, thresholds, evaluation windows, and actions."),
        ("spanner.googleapis.com/Instance", "Azure Cosmos DB / Azure SQL", "Requires workload-level architecture assessment; no direct universal equivalent."),
        ("dataproc.googleapis.com/Cluster", "Azure Databricks / HDInsight", "Collect node types, autoscaling, Spark configuration, jobs, and storage."),
        ("dataflow.googleapis.com/Job", "Azure Data Factory / Stream Analytics / Databricks", "Map batch/stream semantics, throughput, windows, and state."),
        ("aiplatform.googleapis.com/", "Azure Machine Learning", "Collect model endpoints, GPUs, pipelines, registries, and data dependencies."),
        ("backupdr.googleapis.com/", "Azure Backup / Azure Site Recovery", "Map protected workloads, retention, RPO/RTO, backup frequency, and immutability."),
        ("netapp.googleapis.com/", "Azure NetApp Files", "Map protocol, capacity, service level, snapshots, and export/security policy."),
        ("osconfig.googleapis.com/", "Azure Update Manager / Azure Automation / Azure Arc", "Recreate patch schedules, OS policies, reboot behavior, and instance targeting."),
    ]
    for prefix, target, note in mapping:
        if asset_type.startswith(prefix):
            return target, note
    return "Manual architecture assessment", "No one-to-one mapping is encoded; review service capabilities and workload requirements."


def azure_vm_family(vcpus: Optional[int], memory_gib: Optional[float], gpu_count: int, machine_type: str) -> Tuple[str, str]:
    if gpu_count > 0:
        return "Azure N-series", "GPU model, GPU memory, interconnect, driver, and quota must be matched explicitly."

    machine_type = (machine_type or "").lower()
    if machine_type.startswith("t2a"):
        return "Azure ARM-based Dps/Eps family or x64 fallback", "Validate application and container image compatibility with Arm64."

    if not vcpus or memory_gib is None:
        return "Manual selection", "Machine-type specifications could not be retrieved."

    ratio = memory_gib / max(vcpus, 1)
    if memory_gib >= 512 or ratio >= 12:
        family = "Azure M-series (memory optimized)"
    elif ratio >= 6:
        family = "Azure E-series (memory optimized)"
    elif ratio <= 2.5:
        family = "Azure F-series (compute optimized)"
    else:
        family = "Azure D-series (general purpose)"

    note = f"Select an available SKU with at least {vcpus} vCPU and {memory_gib:g} GiB RAM in the target Azure region."
    return family, note


# -----------------------------------------------------------------------------
# Compute Engine inventory
# -----------------------------------------------------------------------------

def list_disks(project: str, compute, errors: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    by_self_link: Dict[str, Dict[str, Any]] = {}
    token = None

    while True:
        try:
            kwargs: Dict[str, Any] = {"project": project, "maxResults": 500}
            if token:
                kwargs["pageToken"] = token
            response = execute(compute.disks().aggregatedList(**kwargs))
        except Exception as ex:
            errors.append({"Area": "Compute disks", "Project": project, "Error": str(ex)})
            break

        for scope_name, scoped in response.get("items", {}).items():
            for disk in scoped.get("disks", []) or []:
                zone = basename(disk.get("zone", ""))
                region = basename(disk.get("region", "")) or zone_to_region(zone)
                users = disk.get("users", []) or []
                row = {
                    "Project": project,
                    "Disk": disk.get("name", ""),
                    "Scope": "Regional" if disk.get("region") else "Zonal",
                    "Region": region,
                    "Zone": zone,
                    "Size GiB": int(disk.get("sizeGb", 0) or 0),
                    "Disk Type": basename(disk.get("type", "")),
                    "Status": disk.get("status", ""),
                    "Attached To": join_values(basename(user) for user in users),
                    "Attached Count": len(users),
                    "Source Image": disk.get("sourceImage", ""),
                    "Source Snapshot": disk.get("sourceSnapshot", ""),
                    "Licenses": join_values(disk.get("licenses", []) or []),
                    "Provisioned IOPS": disk.get("provisionedIops", ""),
                    "Provisioned Throughput": disk.get("provisionedThroughput", ""),
                    "Physical Block Size Bytes": disk.get("physicalBlockSizeBytes", ""),
                    "Architecture": disk.get("architecture", ""),
                    "Labels": key_value_string(disk.get("labels", {})),
                    "Self Link": disk.get("selfLink", ""),
                    "Azure Target": "Azure Managed Disk",
                }
                rows.append(row)
                if disk.get("selfLink"):
                    by_self_link[disk["selfLink"]] = row

        token = response.get("nextPageToken")
        if not token:
            break

    return rows, by_self_link


def get_machine_type_specs(project: str, zone: str, machine_type: str, compute,
                           cache: Dict[Tuple[str, str, str], Dict[str, Any]],
                           errors: List[Dict[str, str]]) -> Dict[str, Any]:
    key = (project, zone, machine_type)
    if key in cache:
        return cache[key]

    try:
        value = execute(compute.machineTypes().get(
            project=project,
            zone=zone,
            machineType=machine_type,
        ))
        result = {
            "vCPUs": int(value.get("guestCpus", 0) or 0),
            "Memory GiB": round(float(value.get("memoryMb", 0) or 0) / 1024.0, 2),
            "Shared CPU": bool(value.get("isSharedCpu", False)),
            "Maximum Persistent Disks": value.get("maximumPersistentDisks", ""),
            "Maximum Persistent Disk Size Gb": value.get("maximumPersistentDisksSizeGb", ""),
        }
    except Exception as ex:
        vcpus, memory_gib = parse_custom_machine_type(machine_type)
        result = {
            "vCPUs": vcpus,
            "Memory GiB": memory_gib,
            "Shared CPU": "",
            "Maximum Persistent Disks": "",
            "Maximum Persistent Disk Size Gb": "",
        }
        errors.append({
            "Area": "Machine type lookup",
            "Project": project,
            "Error": f"{zone}/{machine_type}: {ex}",
        })

    cache[key] = result
    return result


def list_instances(project: str, compute, disk_lookup: Dict[str, Dict[str, Any]],
                   machine_cache: Dict[Tuple[str, str, str], Dict[str, Any]],
                   errors: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], set[str]]:
    vm_rows: List[Dict[str, Any]] = []
    nic_rows: List[Dict[str, Any]] = []
    regions: set[str] = set()
    token = None

    while True:
        try:
            kwargs: Dict[str, Any] = {"project": project, "maxResults": 500}
            if token:
                kwargs["pageToken"] = token
            response = execute(compute.instances().aggregatedList(**kwargs))
        except Exception as ex:
            errors.append({"Area": "Compute instances", "Project": project, "Error": str(ex)})
            break

        for _, scoped in response.get("items", {}).items():
            for inst in scoped.get("instances", []) or []:
                zone = basename(inst.get("zone", ""))
                region = zone_to_region(zone)
                if region:
                    regions.add(region)
                machine_type = basename(inst.get("machineType", ""))
                specs = get_machine_type_specs(project, zone, machine_type, compute, machine_cache, errors)

                metadata = metadata_to_dict(inst)
                labels = inst.get("labels", {}) or {}
                tags = inst.get("tags", {}).get("items", []) or []
                scheduling = inst.get("scheduling", {}) or {}
                provisioning_model = scheduling.get("provisioningModel", "STANDARD")

                disk_details: List[str] = []
                total_disk_gib = 0
                boot_disk_gib = 0
                persistent_disk_count = 0
                local_ssd_count = 0
                disk_types: List[str] = []
                source_images: List[str] = []
                licenses: List[str] = []

                for attached in inst.get("disks", []) or []:
                    source = attached.get("source", "")
                    disk = disk_lookup.get(source, {})
                    if attached.get("type") == "SCRATCH":
                        local_ssd_count += 1
                        disk_details.append(f"{attached.get('deviceName', '')}:local-ssd")
                        continue
                    persistent_disk_count += 1
                    size = int(disk.get("Size GiB", 0) or 0)
                    total_disk_gib += size
                    if attached.get("boot"):
                        boot_disk_gib = size
                    dtype = disk.get("Disk Type", "")
                    if dtype:
                        disk_types.append(dtype)
                    if disk.get("Source Image"):
                        source_images.append(disk["Source Image"])
                    if disk.get("Licenses"):
                        licenses.extend(disk["Licenses"].split(";"))
                    disk_details.append(
                        f"{attached.get('deviceName', basename(source))}:{size}GiB:{dtype}:"
                        f"boot={bool(attached.get('boot'))}:mode={attached.get('mode', '')}"
                    )

                internal_ips: List[str] = []
                external_ips: List[str] = []
                networks: List[str] = []
                subnets: List[str] = []
                network_tiers: List[str] = []

                for nic in inst.get("networkInterfaces", []) or []:
                    internal_ip = nic.get("networkIP", "")
                    access_configs = nic.get("accessConfigs", []) or []
                    nic_external_ips = [cfg.get("natIP", "") for cfg in access_configs if cfg.get("natIP")]
                    tiers = [cfg.get("networkTier", "") for cfg in access_configs if cfg.get("networkTier")]
                    internal_ips.extend([internal_ip] if internal_ip else [])
                    external_ips.extend(nic_external_ips)
                    networks.append(basename(nic.get("network", "")))
                    subnets.append(basename(nic.get("subnetwork", "")))
                    network_tiers.extend(tiers)
                    nic_rows.append({
                        "Project": project,
                        "VM": inst.get("name", ""),
                        "Region": region,
                        "Zone": zone,
                        "NIC": nic.get("name", ""),
                        "Network": basename(nic.get("network", "")),
                        "Subnetwork": basename(nic.get("subnetwork", "")),
                        "Internal IP": internal_ip,
                        "External IPs": join_values(nic_external_ips),
                        "Network Tiers": join_values(tiers),
                        "IP Stack Type": nic.get("stackType", ""),
                        "IPv6 Address": nic.get("ipv6Address", ""),
                        "Queue Count": nic.get("queueCount", ""),
                        "NIC Type": nic.get("nicType", ""),
                    })

                gpu_count = sum(int(gpu.get("acceleratorCount", 0) or 0) for gpu in inst.get("guestAccelerators", []) or [])
                gpu_types = [basename(gpu.get("acceleratorType", "")) for gpu in inst.get("guestAccelerators", []) or []]
                os_name = infer_os(join_values(source_images), licenses, metadata)
                azure_family, azure_note = azure_vm_family(
                    specs.get("vCPUs"), specs.get("Memory GiB"), gpu_count, machine_type
                )

                vm_rows.append({
                    "Project": project,
                    "VM": inst.get("name", ""),
                    "Instance ID": inst.get("id", ""),
                    "Region": region,
                    "Zone": zone,
                    "Status": inst.get("status", ""),
                    "Creation Timestamp": inst.get("creationTimestamp", ""),
                    "Machine Type": machine_type,
                    "vCPUs": specs.get("vCPUs"),
                    "Memory GiB": specs.get("Memory GiB"),
                    "Shared CPU": specs.get("Shared CPU"),
                    "CPU Platform": inst.get("cpuPlatform", ""),
                    "Minimum CPU Platform": inst.get("minCpuPlatform", ""),
                    "OS Guess": os_name,
                    "Provisioning Model": provisioning_model,
                    "Preemptible": bool(scheduling.get("preemptible", False)),
                    "Automatic Restart": scheduling.get("automaticRestart", ""),
                    "On Host Maintenance": scheduling.get("onHostMaintenance", ""),
                    "Availability Domain": scheduling.get("nodeAffinities", ""),
                    "Deletion Protection": bool(inst.get("deletionProtection", False)),
                    "Confidential Compute": bool(inst.get("confidentialInstanceConfig", {}).get("enableConfidentialCompute", False)),
                    "Shielded Secure Boot": bool(inst.get("shieldedInstanceConfig", {}).get("enableSecureBoot", False)),
                    "Shielded vTPM": bool(inst.get("shieldedInstanceConfig", {}).get("enableVtpm", False)),
                    "Shielded Integrity Monitoring": bool(inst.get("shieldedInstanceConfig", {}).get("enableIntegrityMonitoring", False)),
                    "GPU Count": gpu_count,
                    "GPU Types": join_values(gpu_types),
                    "Persistent Disk Count": persistent_disk_count,
                    "Local SSD Count": local_ssd_count,
                    "Total Persistent Disk GiB": total_disk_gib,
                    "Boot Disk GiB": boot_disk_gib,
                    "Disk Types": join_values(sorted(set(disk_types))),
                    "Disk Details": join_values(disk_details),
                    "NIC Count": len(inst.get("networkInterfaces", []) or []),
                    "Networks": join_values(sorted(set(networks))),
                    "Subnetworks": join_values(sorted(set(subnets))),
                    "Internal IPs": join_values(internal_ips),
                    "External IPs": join_values(external_ips),
                    "External IP Count": len(external_ips),
                    "Network Tiers": join_values(sorted(set(network_tiers))),
                    "Can IP Forward": bool(inst.get("canIpForward", False)),
                    "Service Accounts": join_values(sa.get("email", "") for sa in inst.get("serviceAccounts", []) or []),
                    "Tags": join_values(tags),
                    "Labels": key_value_string(labels),
                    "Metadata Keys": join_values(sorted(metadata.keys())),
                    "GKE Cluster": labels.get("goog-k8s-cluster-name", ""),
                    "GKE Node Pool": labels.get("goog-k8s-node-pool-name", labels.get("goog-gke-node-pool", "")),
                    "Managed Instance Group": "",
                    "MIG Scope": "",
                    "MIG Current Action": "",
                    "Workload Association": "GKE Node" if labels.get("goog-k8s-cluster-name") else "Standalone VM",
                    "Azure Compute Target": "AKS Node Pool" if labels.get("goog-k8s-cluster-name") else "Azure Virtual Machine",
                    "Azure VM Family Candidate": azure_family,
                    "Azure Sizing Note": azure_note,
                    "Self Link": inst.get("selfLink", ""),
                })

        token = response.get("nextPageToken")
        if not token:
            break

    return vm_rows, nic_rows, regions


def list_managed_instance_groups(project: str, compute, regions: set[str],
                                 errors: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], Dict[Tuple[str, str, str], Dict[str, str]]]:
    mig_rows: List[Dict[str, Any]] = []
    vm_map: Dict[Tuple[str, str, str], Dict[str, str]] = {}

    # Zonal MIGs.
    token = None
    while True:
        try:
            kwargs: Dict[str, Any] = {"project": project, "maxResults": 500}
            if token:
                kwargs["pageToken"] = token
            response = execute(compute.instanceGroupManagers().aggregatedList(**kwargs))
        except Exception as ex:
            errors.append({"Area": "Zonal MIG inventory", "Project": project, "Error": str(ex)})
            break

        for _, scoped in response.get("items", {}).items():
            for mig in scoped.get("instanceGroupManagers", []) or []:
                zone = basename(mig.get("zone", ""))
                name = mig.get("name", "")
                members: List[Dict[str, Any]] = []
                try:
                    member_response = execute(compute.instanceGroupManagers().listManagedInstances(
                        project=project, zone=zone, instanceGroupManager=name
                    ))
                    members = member_response.get("managedInstances", []) or []
                except Exception as ex:
                    errors.append({"Area": "Zonal MIG members", "Project": project, "Error": f"{zone}/{name}: {ex}"})

                for member in members:
                    vm_url = member.get("instance", "")
                    vm_zone = extract_path_value(vm_url, "zones") or zone
                    vm_name = extract_path_value(vm_url, "instances") or basename(vm_url)
                    vm_map[(project, vm_zone, vm_name)] = {
                        "Managed Instance Group": name,
                        "MIG Scope": "Zonal",
                        "MIG Current Action": member.get("currentAction", ""),
                    }

                mig_rows.append({
                    "Project": project,
                    "MIG": name,
                    "Scope": "Zonal",
                    "Region": zone_to_region(zone),
                    "Zone": zone,
                    "Target Size": mig.get("targetSize", ""),
                    "Current Members": len(members),
                    "Instance Template": basename(mig.get("instanceTemplate", "")),
                    "Base Instance Name": mig.get("baseInstanceName", ""),
                    "Status Stable": mig.get("status", {}).get("isStable", ""),
                    "Distribution Policy": compact_json(mig.get("distributionPolicy", {})),
                    "Update Policy": compact_json(mig.get("updatePolicy", {})),
                    "Autohealing Policies": compact_json(mig.get("autoHealingPolicies", [])),
                    "Azure Target": "Azure Virtual Machine Scale Sets",
                })

        token = response.get("nextPageToken")
        if not token:
            break

    # Regional MIGs. Query only regions containing discovered VMs to avoid dozens
    # of unnecessary API calls. A zero-instance MIG in an otherwise unused region
    # may therefore not appear and will be recorded in the recommendations sheet.
    for region in sorted(regions):
        token = None
        while True:
            try:
                kwargs = {"project": project, "region": region, "maxResults": 500}
                if token:
                    kwargs["pageToken"] = token
                response = execute(compute.regionInstanceGroupManagers().list(**kwargs))
            except Exception as ex:
                # API-disabled/permission errors can repeat by region; retain one row.
                errors.append({"Area": "Regional MIG inventory", "Project": project, "Error": f"{region}: {ex}"})
                break

            for mig in response.get("instanceGroupManagers", []) or []:
                name = mig.get("name", "")
                members: List[Dict[str, Any]] = []
                try:
                    member_response = execute(compute.regionInstanceGroupManagers().listManagedInstances(
                        project=project, region=region, instanceGroupManager=name
                    ))
                    members = member_response.get("managedInstances", []) or []
                except Exception as ex:
                    errors.append({"Area": "Regional MIG members", "Project": project, "Error": f"{region}/{name}: {ex}"})

                for member in members:
                    vm_url = member.get("instance", "")
                    vm_zone = extract_path_value(vm_url, "zones")
                    vm_name = extract_path_value(vm_url, "instances") or basename(vm_url)
                    vm_map[(project, vm_zone, vm_name)] = {
                        "Managed Instance Group": name,
                        "MIG Scope": "Regional",
                        "MIG Current Action": member.get("currentAction", ""),
                    }

                mig_rows.append({
                    "Project": project,
                    "MIG": name,
                    "Scope": "Regional",
                    "Region": region,
                    "Zone": "",
                    "Target Size": mig.get("targetSize", ""),
                    "Current Members": len(members),
                    "Instance Template": basename(mig.get("instanceTemplate", "")),
                    "Base Instance Name": mig.get("baseInstanceName", ""),
                    "Status Stable": mig.get("status", {}).get("isStable", ""),
                    "Distribution Policy": compact_json(mig.get("distributionPolicy", {})),
                    "Update Policy": compact_json(mig.get("updatePolicy", {})),
                    "Autohealing Policies": compact_json(mig.get("autoHealingPolicies", [])),
                    "Azure Target": "Azure Virtual Machine Scale Sets",
                })

            token = response.get("nextPageToken")
            if not token:
                break

    return mig_rows, vm_map


# -----------------------------------------------------------------------------
# GKE inventory and VM-to-node-pool association
# -----------------------------------------------------------------------------

def list_instance_group_members(project: str, group_url: str, compute,
                                errors: List[Dict[str, str]]) -> List[Tuple[str, str, str]]:
    zone = extract_path_value(group_url, "zones")
    group_name = extract_path_value(group_url, "instanceGroups") or extract_path_value(group_url, "instanceGroupManagers") or basename(group_url)
    if not zone or not group_name:
        return []

    members: List[Tuple[str, str, str]] = []
    token = None
    while True:
        try:
            kwargs: Dict[str, Any] = {
                "project": project,
                "zone": zone,
                "instanceGroup": group_name,
                "body": {"instanceState": "ALL"},
                "maxResults": 500,
            }
            if token:
                kwargs["pageToken"] = token
            response = execute(compute.instanceGroups().listInstances(**kwargs))
            for item in response.get("items", []) or []:
                vm_url = item.get("instance", "")
                vm_zone = extract_path_value(vm_url, "zones") or zone
                vm_name = extract_path_value(vm_url, "instances") or basename(vm_url)
                members.append((project, vm_zone, vm_name))
            token = response.get("nextPageToken")
            if not token:
                break
        except HttpError as ex:
            # GKE (especially Autopilot, "gk3-" clusters) reports instance group
            # URLs that can be stale or transient. The backing group may already
            # be deleted by the time we query it, yielding a benign 404. Record it
            # as an informational skip rather than a hard error.
            if getattr(ex, "resp", None) is not None and ex.resp.status == 404:
                errors.append({"Area": "GKE node group members (skipped, transient/not found)",
                               "Project": project, "Error": f"{zone}/{group_name}: instance group no longer exists (404)"})
            else:
                errors.append({"Area": "GKE node group members", "Project": project, "Error": f"{zone}/{group_name}: {ex}"})
            break
        except Exception as ex:
            errors.append({"Area": "GKE node group members", "Project": project, "Error": f"{zone}/{group_name}: {ex}"})
            break
    return members


def list_gke(project: str, container, compute, errors: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[Tuple[str, str, str], Dict[str, str]]]:
    cluster_rows: List[Dict[str, Any]] = []
    nodepool_rows: List[Dict[str, Any]] = []
    vm_map: Dict[Tuple[str, str, str], Dict[str, str]] = {}

    try:
        response = execute(container.projects().locations().clusters().list(
            parent=f"projects/{project}/locations/-"
        ))
    except Exception as ex:
        errors.append({"Area": "GKE clusters", "Project": project, "Error": str(ex)})
        return cluster_rows, nodepool_rows, vm_map

    for cluster in response.get("clusters", []) or []:
        cluster_name = cluster.get("name", "")
        location = cluster.get("location", cluster.get("zone", ""))
        autopilot = bool(cluster.get("autopilot", {}).get("enabled", False))
        private_config = cluster.get("privateClusterConfig", {}) or {}
        release_channel = cluster.get("releaseChannel", {}).get("channel", "")

        cluster_rows.append({
            "Project": project,
            "Cluster": cluster_name,
            "Location": location,
            "Mode": "Autopilot" if autopilot else "Standard",
            "Status": cluster.get("status", ""),
            "Current Master Version": cluster.get("currentMasterVersion", ""),
            "Current Node Version": cluster.get("currentNodeVersion", ""),
            "Current Node Count": cluster.get("currentNodeCount", ""),
            "Initial Cluster Version": cluster.get("initialClusterVersion", ""),
            "Network": basename(cluster.get("network", "")),
            "Subnetwork": basename(cluster.get("subnetwork", "")),
            "Cluster IPv4 CIDR": cluster.get("clusterIpv4Cidr", ""),
            "Services IPv4 CIDR": cluster.get("servicesIpv4Cidr", ""),
            "Private Nodes": private_config.get("enablePrivateNodes", ""),
            "Private Endpoint": private_config.get("enablePrivateEndpoint", ""),
            "Master IPv4 CIDR": private_config.get("masterIpv4CidrBlock", ""),
            "Release Channel": release_channel,
            "Workload Identity Pool": cluster.get("workloadIdentityConfig", {}).get("workloadPool", ""),
            "Logging Service": cluster.get("loggingService", ""),
            "Monitoring Service": cluster.get("monitoringService", ""),
            "Database Encryption": compact_json(cluster.get("databaseEncryption", {})),
            "Resource Labels": key_value_string(cluster.get("resourceLabels", {})),
            "Azure Target": "Azure Kubernetes Service (AKS)",
        })

        for nodepool in cluster.get("nodePools", []) or []:
            config = nodepool.get("config", {}) or {}
            autoscaling = nodepool.get("autoscaling", {}) or {}
            group_urls = (
                nodepool.get("instanceGroupUrls", [])
                or nodepool.get("managedInstanceGroupUrls", [])
                or []
            )
            member_keys: List[Tuple[str, str, str]] = []
            for group_url in group_urls:
                member_keys.extend(list_instance_group_members(project, group_url, compute, errors))

            pool_name = nodepool.get("name", "")
            for key in member_keys:
                vm_map[key] = {
                    "GKE Cluster": cluster_name,
                    "GKE Node Pool": pool_name,
                    "Workload Association": "GKE Node",
                    "Azure Compute Target": "AKS Node Pool",
                }

            accelerators = config.get("accelerators", []) or []
            gpu_count_per_node = sum(int(a.get("acceleratorCount", 0) or 0) for a in accelerators)
            gpu_types = [a.get("acceleratorType", "") for a in accelerators]
            machine_type = config.get("machineType", "")

            nodepool_rows.append({
                "Project": project,
                "Cluster": cluster_name,
                "Cluster Location": location,
                "Cluster Mode": "Autopilot" if autopilot else "Standard",
                "Node Pool": pool_name,
                "Status": nodepool.get("status", ""),
                "Version": nodepool.get("version", ""),
                "Machine Type": machine_type,
                "Image Type": config.get("imageType", ""),
                "Disk Size GiB": config.get("diskSizeGb", ""),
                "Disk Type": config.get("diskType", ""),
                "Local SSD Count": config.get("localSsdCount", ""),
                "Boot Disk KMS Key": config.get("bootDiskKmsKey", ""),
                "Service Account": config.get("serviceAccount", ""),
                "Spot": config.get("spot", ""),
                "Preemptible": config.get("preemptible", ""),
                "Initial Node Count": nodepool.get("initialNodeCount", ""),
                "Discovered VM Count": len(set(member_keys)),
                "Locations": join_values(nodepool.get("locations", []) or []),
                "Autoscaling Enabled": autoscaling.get("enabled", ""),
                "Autoscaling Min Nodes": autoscaling.get("minNodeCount", ""),
                "Autoscaling Max Nodes": autoscaling.get("maxNodeCount", ""),
                "Autoscaling Total Min Nodes": autoscaling.get("totalMinNodeCount", ""),
                "Autoscaling Total Max Nodes": autoscaling.get("totalMaxNodeCount", ""),
                "Max Pods Per Node": nodepool.get("maxPodsConstraint", {}).get("maxPodsPerNode", ""),
                "GPU Count per Node": gpu_count_per_node,
                "GPU Types": join_values(gpu_types),
                "Labels": key_value_string(config.get("labels", {})),
                "Resource Labels": key_value_string(nodepool.get("resourceLabels", {})),
                "Taints": compact_json(config.get("taints", [])),
                "Tags": join_values(config.get("tags", []) or []),
                "OAuth Scopes": join_values(config.get("oauthScopes", []) or []),
                "Instance Groups": join_values(group_urls),
                "Auto Upgrade": nodepool.get("management", {}).get("autoUpgrade", ""),
                "Auto Repair": nodepool.get("management", {}).get("autoRepair", ""),
                "Upgrade Strategy": nodepool.get("upgradeSettings", {}).get("strategy", ""),
                "Max Surge": nodepool.get("upgradeSettings", {}).get("maxSurge", ""),
                "Max Unavailable": nodepool.get("upgradeSettings", {}).get("maxUnavailable", ""),
                "Shielded Secure Boot": config.get("shieldedInstanceConfig", {}).get("enableSecureBoot", ""),
                "Shielded Integrity Monitoring": config.get("shieldedInstanceConfig", {}).get("enableIntegrityMonitoring", ""),
                "GVNIC Enabled": config.get("gvnic", {}).get("enabled", ""),
                "Azure Target": "AKS Node Pool",
                "Azure Sizing Note": "Match node count/autoscaling plus pod CPU-memory requests, limits, daemonsets, system reserve, and availability-zone design.",
            })

    return cluster_rows, nodepool_rows, vm_map


def apply_associations(vm_rows: List[Dict[str, Any]], mig_map: Dict[Tuple[str, str, str], Dict[str, str]],
                       gke_map: Dict[Tuple[str, str, str], Dict[str, str]]) -> None:
    for row in vm_rows:
        key = (row.get("Project", ""), row.get("Zone", ""), row.get("VM", ""))
        if key in mig_map:
            row.update(mig_map[key])
            if row.get("Workload Association") != "GKE Node":
                row["Workload Association"] = "Managed Instance Group VM"
                row["Azure Compute Target"] = "Azure Virtual Machine Scale Set"
        if key in gke_map:
            row.update(gke_map[key])


# -----------------------------------------------------------------------------
# Cloud SQL inventory
# -----------------------------------------------------------------------------

def cloud_sql_engine(database_version: str) -> Tuple[str, str, str]:
    """Return (engine, Azure target, migration note) from a Cloud SQL databaseVersion."""
    version = (database_version or "").upper()
    if version.startswith("MYSQL"):
        return ("MySQL", "Azure Database for MySQL Flexible Server",
                "Match engine version, vCPU/RAM tier, storage/IOPS, HA, read replicas, and server flags.")
    if version.startswith("POSTGRES"):
        return ("PostgreSQL", "Azure Database for PostgreSQL Flexible Server",
                "Match engine version, vCPU/RAM tier, storage/IOPS, HA, read replicas, extensions, and server flags.")
    if version.startswith("SQLSERVER"):
        return ("SQL Server", "Azure SQL Managed Instance / SQL Server on Azure VM",
                "Match edition, version, vCPU/RAM, storage/IOPS, HA/Always On, licensing, and feature compatibility.")
    return ("Unknown", "Manual assessment", "Unrecognized Cloud SQL engine; review the target service manually.")


def parse_cloud_sql_tier(tier: str) -> Tuple[Optional[int], Optional[float]]:
    """Derive (vCPU, memory GiB) from a Cloud SQL machine tier where possible."""
    tier = (tier or "").lower()
    # Custom tiers: db-custom-<vCPU>-<memoryMB>.
    match = re.match(r"db-custom-(\d+)-(\d+)", tier)
    if match:
        return int(match.group(1)), round(int(match.group(2)) / 1024.0, 2)
    # Legacy first/second-gen N1 tiers.
    match = re.match(r"db-n1-(standard|highmem)-(\d+)", tier)
    if match:
        vcpus = int(match.group(2))
        memory_gib = vcpus * (3.75 if match.group(1) == "standard" else 6.5)
        return vcpus, round(memory_gib, 2)
    # Shared-core tiers.
    shared = {"db-f1-micro": (None, 0.6), "db-g1-small": (None, 1.7)}
    if tier in shared:
        return shared[tier]
    return None, None


def list_cloud_sql(project: str, sqladmin, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    try:
        response = execute(sqladmin.instances().list(project=project))
    except HttpError as ex:
        status = getattr(getattr(ex, "resp", None), "status", None)
        if status in (403, 404):
            # API disabled or no access on this project; record once and continue.
            errors.append({"Area": "Cloud SQL (skipped, API disabled or access denied)",
                           "Project": project, "Error": str(ex)})
        else:
            errors.append({"Area": "Cloud SQL instances", "Project": project, "Error": str(ex)})
        return rows
    except Exception as ex:
        errors.append({"Area": "Cloud SQL instances", "Project": project, "Error": str(ex)})
        return rows

    for inst in response.get("items", []) or []:
        name = inst.get("name", "")
        settings = inst.get("settings", {}) or {}
        backup = settings.get("backupConfiguration", {}) or {}
        backup_retention = backup.get("backupRetentionSettings", {}) or {}
        ip_config = settings.get("ipConfiguration", {}) or {}
        maintenance = settings.get("maintenanceWindow", {}) or {}
        db_version = inst.get("databaseVersion", "")
        engine, azure_target, azure_note = cloud_sql_engine(db_version)
        tier = settings.get("tier", "")
        vcpus, memory_gib = parse_cloud_sql_tier(tier)

        ip_addresses = inst.get("ipAddresses", []) or []
        public_ips = [ip.get("ipAddress", "") for ip in ip_addresses if ip.get("type") == "PRIMARY"]
        private_ips = [ip.get("ipAddress", "") for ip in ip_addresses if ip.get("type") == "PRIVATE"]

        flags = settings.get("databaseFlags", []) or []
        flag_str = join_values(f"{f.get('name')}={f.get('value')}" for f in flags)
        authorized = ip_config.get("authorizedNetworks", []) or []
        replica_names = inst.get("replicaNames", []) or []

        # Enumerate the individual databases (schemas) hosted on the instance.
        database_names: List[str] = []
        try:
            db_response = execute(sqladmin.databases().list(project=project, instance=name))
            database_names = [db.get("name", "") for db in db_response.get("items", []) or [] if db.get("name")]
        except Exception as ex:
            errors.append({"Area": "Cloud SQL databases", "Project": project, "Error": f"{name}: {ex}"})

        rows.append({
            "Project": project,
            "Instance": name,
            "Engine": engine,
            "Database Version": db_version,
            "Edition": settings.get("edition", ""),
            "Instance Role": inst.get("instanceType", ""),
            "State": inst.get("state", ""),
            "Region": inst.get("region", ""),
            "Primary Zone": inst.get("gceZone", ""),
            "Secondary Zone": inst.get("secondaryGceZone", ""),
            "Availability Type": settings.get("availabilityType", ""),
            "High Availability": "Yes" if settings.get("availabilityType") == "REGIONAL" else "No",
            "Tier": tier,
            "vCPUs": vcpus,
            "Memory GiB": memory_gib,
            "Data Disk Size GiB": settings.get("dataDiskSizeGb", ""),
            "Data Disk Type": settings.get("dataDiskType", ""),
            "Storage Auto Resize": settings.get("storageAutoResize", ""),
            "Storage Auto Resize Limit GiB": settings.get("storageAutoResizeLimit", ""),
            "Provisioned IOPS": settings.get("dataDiskProvisionedIops", ""),
            "Provisioned Throughput": settings.get("dataDiskProvisionedThroughput", ""),
            "Activation Policy": settings.get("activationPolicy", ""),
            "Pricing Plan": settings.get("pricingPlan", ""),
            "Connection Name": inst.get("connectionName", ""),
            "Public IP Enabled": ip_config.get("ipv4Enabled", ""),
            "Private Network": basename(ip_config.get("privateNetwork", "")),
            "SSL Mode": ip_config.get("sslMode", ""),
            "Require SSL": ip_config.get("requireSsl", ""),
            "Authorized Network Count": len(authorized),
            "Public IPs": join_values(public_ips),
            "Private IPs": join_values(private_ips),
            "Master Instance": inst.get("masterInstanceName", ""),
            "Read Replicas": join_values(basename(r) for r in replica_names),
            "Read Replica Count": len(replica_names),
            "Failover Replica": (inst.get("failoverReplica", {}) or {}).get("name", ""),
            "Backups Enabled": backup.get("enabled", ""),
            "Backup Start Time": backup.get("startTime", ""),
            "Point In Time Recovery": backup.get("pointInTimeRecoveryEnabled", ""),
            "Transaction Log Retention Days": backup.get("transactionLogRetentionDays", ""),
            "Retained Backups": backup_retention.get("retainedBackups", ""),
            "Backup Retention Unit": backup_retention.get("retentionUnit", ""),
            "Deletion Protection": settings.get("deletionProtectionEnabled", ""),
            "CMEK Key": basename((inst.get("diskEncryptionConfiguration", {}) or {}).get("kmsKeyName", "")),
            "Maintenance Day": maintenance.get("day", ""),
            "Maintenance Hour": maintenance.get("hour", ""),
            "Maintenance Track": maintenance.get("updateTrack", ""),
            "Database Flags": flag_str,
            "Database Count": len(database_names),
            "Databases": join_values(database_names),
            "Service Account": inst.get("serviceAccountEmailAddress", ""),
            "Maintenance Version": inst.get("maintenanceVersion", ""),
            "Create Time": inst.get("createTime", ""),
            "Labels": key_value_string(settings.get("userLabels", {})),
            "Azure Target": azure_target,
            "Azure Sizing Note": azure_note,
        })

    return rows


# -----------------------------------------------------------------------------
# Cloud Storage inventory
# -----------------------------------------------------------------------------

def list_storage_buckets(project: str, storage, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    token = None

    while True:
        try:
            kwargs: Dict[str, Any] = {"project": project, "projection": "full", "maxResults": 1000}
            if token:
                kwargs["pageToken"] = token
            response = execute(storage.buckets().list(**kwargs))
        except HttpError as ex:
            status = getattr(getattr(ex, "resp", None), "status", None)
            if status in (403, 404):
                errors.append({"Area": "Cloud Storage (skipped, API disabled or access denied)",
                               "Project": project, "Error": str(ex)})
            else:
                errors.append({"Area": "Cloud Storage buckets", "Project": project, "Error": str(ex)})
            break
        except Exception as ex:
            errors.append({"Area": "Cloud Storage buckets", "Project": project, "Error": str(ex)})
            break

        for bucket in response.get("items", []) or []:
            iam = bucket.get("iamConfiguration", {}) or {}
            ubla = iam.get("uniformBucketLevelAccess", {}) or {}
            retention = bucket.get("retentionPolicy", {}) or {}
            autoclass = bucket.get("autoclass", {}) or {}
            soft_delete = bucket.get("softDeletePolicy", {}) or {}
            lifecycle_rules = (bucket.get("lifecycle", {}) or {}).get("rule", []) or []
            rows.append({
                "Project": project,
                "Bucket": bucket.get("name", ""),
                "Location": bucket.get("location", ""),
                "Location Type": bucket.get("locationType", ""),
                "Storage Class": bucket.get("storageClass", ""),
                "Public Access Prevention": iam.get("publicAccessPrevention", ""),
                "Uniform Bucket-Level Access": ubla.get("enabled", ""),
                "Versioning Enabled": (bucket.get("versioning", {}) or {}).get("enabled", ""),
                "Lifecycle Rule Count": len(lifecycle_rules),
                "Retention Period (s)": retention.get("retentionPeriod", ""),
                "Retention Locked": retention.get("isLocked", ""),
                "Autoclass Enabled": autoclass.get("enabled", ""),
                "Autoclass Terminal Class": autoclass.get("terminalStorageClass", ""),
                "Soft Delete Retention (s)": soft_delete.get("retentionDurationSeconds", ""),
                "Default KMS Key": basename((bucket.get("encryption", {}) or {}).get("defaultKmsKeyName", "")),
                "Requester Pays": (bucket.get("billing", {}) or {}).get("requesterPays", ""),
                "RPO": bucket.get("rpo", ""),
                "Default Event-Based Hold": bucket.get("defaultEventBasedHold", ""),
                "Log Bucket": (bucket.get("logging", {}) or {}).get("logBucket", ""),
                "Created": bucket.get("timeCreated", ""),
                "Updated": bucket.get("updated", ""),
                "Labels": key_value_string(bucket.get("labels", {})),
                "Azure Target": "Azure Blob Storage / ADLS Gen2",
                "Azure Sizing Note": "Collect stored bytes, object count, and request/egress patterns (via Cloud Monitoring) before selecting tier and redundancy.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    return rows


# -----------------------------------------------------------------------------
# BigQuery inventory
# -----------------------------------------------------------------------------

BQ_MAX_TABLES_PER_DATASET = 1000


def list_bigquery(project: str, bigquery, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    datasets: List[Dict[str, Any]] = []
    token = None

    while True:
        try:
            kwargs: Dict[str, Any] = {"projectId": project, "all": True, "maxResults": 1000}
            if token:
                kwargs["pageToken"] = token
            response = execute(bigquery.datasets().list(**kwargs))
        except HttpError as ex:
            status = getattr(getattr(ex, "resp", None), "status", None)
            if status in (403, 404):
                errors.append({"Area": "BigQuery (skipped, API disabled or access denied)",
                               "Project": project, "Error": str(ex)})
            else:
                errors.append({"Area": "BigQuery datasets", "Project": project, "Error": str(ex)})
            return rows
        except Exception as ex:
            errors.append({"Area": "BigQuery datasets", "Project": project, "Error": str(ex)})
            return rows

        datasets.extend(response.get("datasets", []) or [])
        token = response.get("nextPageToken")
        if not token:
            break

    for ds in datasets:
        dataset_id = (ds.get("datasetReference", {}) or {}).get("datasetId", "")
        if not dataset_id:
            continue

        detail: Dict[str, Any] = {}
        try:
            detail = execute(bigquery.datasets().get(projectId=project, datasetId=dataset_id))
        except Exception as ex:
            errors.append({"Area": "BigQuery dataset detail", "Project": project, "Error": f"{dataset_id}: {ex}"})

        total_bytes = 0
        long_term_bytes = 0
        total_rows = 0
        type_counts: Counter = Counter()
        inspected = 0
        size_complete = True
        table_token = None

        while True:
            try:
                t_kwargs: Dict[str, Any] = {"projectId": project, "datasetId": dataset_id, "maxResults": 1000}
                if table_token:
                    t_kwargs["pageToken"] = table_token
                t_response = execute(bigquery.tables().list(**t_kwargs))
            except Exception as ex:
                errors.append({"Area": "BigQuery tables", "Project": project, "Error": f"{dataset_id}: {ex}"})
                break

            for table in t_response.get("tables", []) or []:
                type_counts[table.get("type", "TABLE")] += 1
                table_id = (table.get("tableReference", {}) or {}).get("tableId", "")
                if inspected < BQ_MAX_TABLES_PER_DATASET and table_id:
                    try:
                        t_detail = execute(bigquery.tables().get(
                            projectId=project, datasetId=dataset_id, tableId=table_id))
                        total_bytes += int(t_detail.get("numBytes", 0) or 0)
                        long_term_bytes += int(t_detail.get("numLongTermBytes", 0) or 0)
                        total_rows += int(t_detail.get("numRows", 0) or 0)
                    except Exception:
                        size_complete = False
                    inspected += 1
                else:
                    size_complete = False

            table_token = t_response.get("nextPageToken")
            if not table_token:
                break

        access_entries = detail.get("access", []) or []
        rows.append({
            "Project": project,
            "Dataset": dataset_id,
            "Location": detail.get("location", ds.get("location", "")),
            "Description": detail.get("description", ""),
            "Table Count": sum(type_counts.values()),
            "Table Types": key_value_string(dict(type_counts)),
            "Total Logical Bytes": total_bytes,
            "Total Logical GiB": round(total_bytes / (1024 ** 3), 3) if total_bytes else 0,
            "Long-Term Bytes": long_term_bytes,
            "Total Rows": total_rows,
            "Size Complete": "Yes" if size_complete else "No (capped or partial)",
            "Default Table Expiration (ms)": detail.get("defaultTableExpirationMs", ""),
            "Default Partition Expiration (ms)": detail.get("defaultPartitionExpirationMs", ""),
            "Default KMS Key": basename((detail.get("defaultEncryptionConfiguration", {}) or {}).get("kmsKeyName", "")),
            "Access Entry Count": len(access_entries),
            "Created (epoch ms)": detail.get("creationTime", ""),
            "Last Modified (epoch ms)": detail.get("lastModifiedTime", ""),
            "Labels": key_value_string(detail.get("labels", ds.get("labels", {})) or {}),
            "Azure Target": "Microsoft Fabric / Azure Synapse / ADLS Gen2",
            "Azure Sizing Note": "Confirm slot/query patterns, pipelines, partitioning/clustering, and governance; logical bytes are a storage baseline only.",
        })

    return rows


# -----------------------------------------------------------------------------
# Cloud Logging inventory
# -----------------------------------------------------------------------------

def list_cloud_logging(project: str, logging_svc, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    # Log routing sinks.
    token = None
    while True:
        try:
            kwargs: Dict[str, Any] = {"parent": f"projects/{project}"}
            if token:
                kwargs["pageToken"] = token
            response = execute(logging_svc.sinks().list(**kwargs))
        except HttpError as ex:
            status = getattr(getattr(ex, "resp", None), "status", None)
            if status in (403, 404):
                errors.append({"Area": "Cloud Logging (skipped, API disabled or access denied)",
                               "Project": project, "Error": str(ex)})
            else:
                errors.append({"Area": "Cloud Logging sinks", "Project": project, "Error": str(ex)})
            break
        except Exception as ex:
            errors.append({"Area": "Cloud Logging sinks", "Project": project, "Error": str(ex)})
            break

        for sink in response.get("sinks", []) or []:
            rows.append({
                "Project": project,
                "Resource Type": "Log Sink",
                "Name": basename(sink.get("name", "")),
                "Location": "global",
                "Destination": sink.get("destination", ""),
                "Filter": sink.get("filter", ""),
                "Disabled/Locked": sink.get("disabled", False),
                "Retention Days": "",
                "Analytics Enabled": "",
                "Default KMS Key": "",
                "Description": sink.get("description", ""),
                "Created": sink.get("createTime", ""),
                "Azure Target": "Azure Monitor / Log Analytics / Event Hubs",
                "Azure Sizing Note": "Recreate routing to Log Analytics/Event Hubs; capture filter, destination, and volume.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    # Log storage buckets.
    token = None
    while True:
        try:
            kwargs = {"parent": f"projects/{project}/locations/-"}
            if token:
                kwargs["pageToken"] = token
            response = execute(logging_svc.projects().locations().buckets().list(**kwargs))
        except Exception as ex:
            errors.append({"Area": "Cloud Logging buckets", "Project": project, "Error": str(ex)})
            break

        for bucket in response.get("buckets", []) or []:
            name = bucket.get("name", "")
            rows.append({
                "Project": project,
                "Resource Type": "Log Bucket",
                "Name": basename(name),
                "Location": extract_path_value(name, "locations"),
                "Destination": "",
                "Filter": "",
                "Disabled/Locked": bucket.get("locked", False),
                "Retention Days": bucket.get("retentionDays", ""),
                "Analytics Enabled": bucket.get("analyticsEnabled", ""),
                "Default KMS Key": basename((bucket.get("cmekSettings", {}) or {}).get("kmsKeyName", "")),
                "Description": bucket.get("description", ""),
                "Created": bucket.get("createTime", ""),
                "Azure Target": "Azure Monitor Log Analytics workspace",
                "Azure Sizing Note": "Map retention, analytics, and CMEK to a Log Analytics workspace; capture ingestion volume.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    return rows


# -----------------------------------------------------------------------------
# Backup and DR Service inventory
# -----------------------------------------------------------------------------

def list_backup_dr(project: str, backupdr, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    parent = f"projects/{project}/locations/-"

    def paged(resource_getter, key: str, area: str) -> List[Dict[str, Any]]:
        results: List[Dict[str, Any]] = []
        try:
            resource = resource_getter()
        except AttributeError as ex:
            errors.append({"Area": f"{area} (not available in API version)", "Project": project, "Error": str(ex)})
            return results

        token = None
        while True:
            try:
                kwargs: Dict[str, Any] = {"parent": parent}
                if token:
                    kwargs["pageToken"] = token
                response = execute(resource.list(**kwargs))
            except HttpError as ex:
                status = getattr(getattr(ex, "resp", None), "status", None)
                if status in (403, 404):
                    errors.append({"Area": f"{area} (skipped, API disabled or access denied)",
                                   "Project": project, "Error": str(ex)})
                else:
                    errors.append({"Area": area, "Project": project, "Error": str(ex)})
                break
            except Exception as ex:
                errors.append({"Area": area, "Project": project, "Error": str(ex)})
                break
            results.extend(response.get(key, []) or [])
            token = response.get("nextPageToken")
            if not token:
                break
        return results

    for server in paged(lambda: backupdr.projects().locations().managementServers(),
                        "managementServers", "Backup and DR management servers"):
        name = server.get("name", "")
        rows.append({
            "Project": project,
            "Resource Type": "Management Server",
            "Name": basename(name),
            "Location": extract_path_value(name, "locations"),
            "State": server.get("state", ""),
            "Server/Plan Type": server.get("type", ""),
            "Backup Vault": "",
            "Retention": "",
            "Stored Bytes": "",
            "Backup Count": "",
            "Protected Resource Type": "",
            "Networks": join_values(n.get("network", "") for n in server.get("networks", []) or []),
            "Created": server.get("createTime", ""),
            "Azure Target": "Azure Backup / Azure Site Recovery",
            "Azure Sizing Note": "Map protected workloads, retention, RPO/RTO, and backup frequency to Recovery Services vault policies.",
        })

    for vault in paged(lambda: backupdr.projects().locations().backupVaults(),
                       "backupVaults", "Backup and DR backup vaults"):
        name = vault.get("name", "")
        rows.append({
            "Project": project,
            "Resource Type": "Backup Vault",
            "Name": basename(name),
            "Location": extract_path_value(name, "locations"),
            "State": vault.get("state", ""),
            "Server/Plan Type": "",
            "Backup Vault": basename(name),
            "Retention": vault.get("backupMinimumEnforcedRetentionDuration", ""),
            "Stored Bytes": vault.get("totalStoredBytes", ""),
            "Backup Count": vault.get("backupCount", ""),
            "Protected Resource Type": "",
            "Networks": "",
            "Created": vault.get("createTime", ""),
            "Azure Target": "Azure Backup vault / Recovery Services vault",
            "Azure Sizing Note": "Map enforced retention, stored bytes, and immutability to Azure Backup vault policies.",
        })

    for plan in paged(lambda: backupdr.projects().locations().backupPlans(),
                      "backupPlans", "Backup and DR backup plans"):
        name = plan.get("name", "")
        rows.append({
            "Project": project,
            "Resource Type": "Backup Plan",
            "Name": basename(name),
            "Location": extract_path_value(name, "locations"),
            "State": plan.get("state", ""),
            "Server/Plan Type": "",
            "Backup Vault": basename(plan.get("backupVault", "")),
            "Retention": "",
            "Stored Bytes": "",
            "Backup Count": "",
            "Protected Resource Type": plan.get("resourceType", ""),
            "Networks": "",
            "Created": plan.get("createTime", ""),
            "Azure Target": "Azure Backup policy",
            "Azure Sizing Note": f"Recreate backup rules ({len(plan.get('backupRules', []) or [])} rule(s)) as Azure Backup policy schedules and retention.",
        })

    return rows


# -----------------------------------------------------------------------------
# NetApp Volumes inventory
# -----------------------------------------------------------------------------

def list_netapp(project: str, netapp, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    parent = f"projects/{project}/locations/-"

    # Storage pools.
    token = None
    while True:
        try:
            kwargs: Dict[str, Any] = {"parent": parent}
            if token:
                kwargs["pageToken"] = token
            response = execute(netapp.projects().locations().storagePools().list(**kwargs))
        except HttpError as ex:
            status = getattr(getattr(ex, "resp", None), "status", None)
            if status in (403, 404):
                errors.append({"Area": "NetApp Volumes (skipped, API disabled or access denied)",
                               "Project": project, "Error": str(ex)})
            else:
                errors.append({"Area": "NetApp storage pools", "Project": project, "Error": str(ex)})
            break
        except Exception as ex:
            errors.append({"Area": "NetApp storage pools", "Project": project, "Error": str(ex)})
            break

        for pool in response.get("storagePools", []) or []:
            name = pool.get("name", "")
            rows.append({
                "Project": project,
                "Resource Type": "Storage Pool",
                "Name": basename(name),
                "Location": extract_path_value(name, "locations"),
                "State": pool.get("state", ""),
                "Service Level": pool.get("serviceLevel", ""),
                "Capacity GiB": pool.get("capacityGib", ""),
                "Allocated Volume GiB": pool.get("volumeCapacityGib", ""),
                "Used GiB": "",
                "Storage Pool": basename(name),
                "Protocols": "",
                "Share Name": "",
                "Network": basename(pool.get("network", "")),
                "Encryption Type": pool.get("encryptionType", ""),
                "LDAP Enabled": pool.get("ldapEnabled", ""),
                "Labels": key_value_string(pool.get("labels", {})),
                "Azure Target": "Azure NetApp Files capacity pool",
                "Azure Sizing Note": "Map service level (Standard/Premium/Ultra) and capacity to an Azure NetApp Files capacity pool.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    # Volumes.
    token = None
    while True:
        try:
            kwargs = {"parent": parent}
            if token:
                kwargs["pageToken"] = token
            response = execute(netapp.projects().locations().volumes().list(**kwargs))
        except Exception as ex:
            errors.append({"Area": "NetApp volumes", "Project": project, "Error": str(ex)})
            break

        for vol in response.get("volumes", []) or []:
            name = vol.get("name", "")
            rows.append({
                "Project": project,
                "Resource Type": "Volume",
                "Name": basename(name),
                "Location": extract_path_value(name, "locations"),
                "State": vol.get("state", ""),
                "Service Level": vol.get("serviceLevel", ""),
                "Capacity GiB": vol.get("capacityGib", ""),
                "Allocated Volume GiB": "",
                "Used GiB": vol.get("usedGib", ""),
                "Storage Pool": basename(vol.get("storagePool", "")),
                "Protocols": join_values(vol.get("protocols", []) or []),
                "Share Name": vol.get("shareName", ""),
                "Network": "",
                "Encryption Type": vol.get("encryptionType", ""),
                "LDAP Enabled": vol.get("ldapEnabled", ""),
                "Labels": key_value_string(vol.get("labels", {})),
                "Azure Target": "Azure NetApp Files volume",
                "Azure Sizing Note": "Map protocol (NFSv3/NFSv4.1/SMB), capacity, service level, snapshots, and export policy to an ANF volume.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    return rows


# -----------------------------------------------------------------------------
# VM Manager (OS Config) inventory
# -----------------------------------------------------------------------------

def list_vm_manager(project: str, osconfig, errors: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    # Patch deployments (project-level).
    token = None
    while True:
        try:
            kwargs: Dict[str, Any] = {"parent": f"projects/{project}"}
            if token:
                kwargs["pageToken"] = token
            response = execute(osconfig.projects().patchDeployments().list(**kwargs))
        except HttpError as ex:
            status = getattr(getattr(ex, "resp", None), "status", None)
            if status in (403, 404):
                errors.append({"Area": "VM Manager (skipped, API disabled or access denied)",
                               "Project": project, "Error": str(ex)})
            else:
                errors.append({"Area": "VM Manager patch deployments", "Project": project, "Error": str(ex)})
            break
        except Exception as ex:
            errors.append({"Area": "VM Manager patch deployments", "Project": project, "Error": str(ex)})
            break

        for item in response.get("patchDeployments", []) or []:
            name = item.get("name", "")
            instance_filter = item.get("instanceFilter", {}) or {}
            patch_config = item.get("patchConfig", {}) or {}
            schedule = "OneTime" if item.get("oneTimeSchedule") else ("Recurring" if item.get("recurringSchedule") else "")
            rows.append({
                "Project": project,
                "Resource Type": "Patch Deployment",
                "Name": basename(name),
                "State": item.get("state", ""),
                "Schedule Type": schedule,
                "Reboot Config": patch_config.get("rebootConfig", ""),
                "Targets All Instances": instance_filter.get("all", ""),
                "Target Zones": join_values(instance_filter.get("zones", []) or []),
                "Target Group Label Sets": len(instance_filter.get("groupLabels", []) or []),
                "Policy Count": "",
                "Duration": item.get("duration", ""),
                "Last Execute Time": item.get("lastExecuteTime", ""),
                "Description": item.get("description", ""),
                "Created": item.get("createTime", ""),
                "Azure Target": "Azure Update Manager / Azure Arc",
                "Azure Sizing Note": "Recreate patch schedules, reboot policy, and instance targeting in Azure Update Manager (with Azure Arc for hybrid).",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    # OS policy assignments (best effort; the location wildcard may not be supported).
    token = None
    while True:
        try:
            kwargs = {"parent": f"projects/{project}/locations/-"}
            if token:
                kwargs["pageToken"] = token
            response = execute(osconfig.projects().locations().osPolicyAssignments().list(**kwargs))
        except Exception as ex:
            errors.append({"Area": "VM Manager OS policy assignments (skipped)", "Project": project, "Error": str(ex)})
            break

        for assignment in response.get("osPolicyAssignments", []) or []:
            name = assignment.get("name", "")
            instance_filter = assignment.get("instanceFilter", {}) or {}
            rows.append({
                "Project": project,
                "Resource Type": "OS Policy Assignment",
                "Name": basename(name),
                "State": assignment.get("rolloutState", ""),
                "Schedule Type": "",
                "Reboot Config": "",
                "Targets All Instances": instance_filter.get("all", ""),
                "Target Zones": "",
                "Target Group Label Sets": len(instance_filter.get("inclusionLabels", []) or []),
                "Policy Count": len(assignment.get("osPolicies", []) or []),
                "Duration": "",
                "Last Execute Time": "",
                "Description": assignment.get("description", ""),
                "Created": assignment.get("revisionCreateTime", ""),
                "Azure Target": "Azure Machine Configuration / Azure Automation",
                "Azure Sizing Note": "Recreate OS policies as Azure Machine Configuration (guest configuration) assignments.",
            })

        token = response.get("nextPageToken")
        if not token:
            break

    return rows


# -----------------------------------------------------------------------------
# Reporting
# -----------------------------------------------------------------------------

def dataframe(rows: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> pd.DataFrame:
    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=columns or [])


def create_asset_summary(assets_df: pd.DataFrame) -> pd.DataFrame:
    if assets_df.empty:
        return pd.DataFrame(columns=["Asset Type", "Count", "Azure Candidate", "Migration Note"])
    summary = assets_df.groupby("Asset Type").size().reset_index(name="Count")
    mappings = summary["Asset Type"].apply(azure_service_mapping)
    summary["Azure Candidate"] = mappings.apply(lambda value: value[0])
    summary["Migration Note"] = mappings.apply(lambda value: value[1])
    return summary.sort_values(["Count", "Asset Type"], ascending=[False, True]).reset_index(drop=True)


def create_azure_baseline(vm_df: pd.DataFrame, disk_df: pd.DataFrame, cluster_df: pd.DataFrame,
                          nodepool_df: pd.DataFrame, mig_df: pd.DataFrame, assets_df: pd.DataFrame,
                          sql_df: pd.DataFrame, storage_df: pd.DataFrame, bigquery_df: pd.DataFrame,
                          logging_df: pd.DataFrame, backupdr_df: pd.DataFrame, netapp_df: pd.DataFrame,
                          vmmanager_df: pd.DataFrame) -> pd.DataFrame:
    def numeric_sum(df: pd.DataFrame, column: str) -> float:
        if df.empty or column not in df.columns:
            return 0
        return float(pd.to_numeric(df[column], errors="coerce").fillna(0).sum())

    running = vm_df[vm_df["Status"] == "RUNNING"] if not vm_df.empty and "Status" in vm_df.columns else pd.DataFrame()
    stopped = vm_df[vm_df["Status"].isin(["TERMINATED", "STOPPED", "SUSPENDED"])] if not vm_df.empty and "Status" in vm_df.columns else pd.DataFrame()
    gke_nodes = vm_df[vm_df["Workload Association"] == "GKE Node"] if not vm_df.empty and "Workload Association" in vm_df.columns else pd.DataFrame()
    mig_vms = vm_df[vm_df["Managed Instance Group"].astype(str) != ""] if not vm_df.empty and "Managed Instance Group" in vm_df.columns else pd.DataFrame()

    sql_primaries = sql_df[sql_df["Instance Role"] == "CLOUD_SQL_INSTANCE"] if not sql_df.empty and "Instance Role" in sql_df.columns else pd.DataFrame()
    sql_replicas = sql_df[sql_df["Instance Role"] == "READ_REPLICA_INSTANCE"] if not sql_df.empty and "Instance Role" in sql_df.columns else pd.DataFrame()
    sql_ha = sql_df[sql_df["High Availability"] == "Yes"] if not sql_df.empty and "High Availability" in sql_df.columns else pd.DataFrame()

    netapp_volumes = netapp_df[netapp_df["Resource Type"] == "Volume"] if not netapp_df.empty and "Resource Type" in netapp_df.columns else pd.DataFrame()
    vm_patch = vmmanager_df[vmmanager_df["Resource Type"] == "Patch Deployment"] if not vmmanager_df.empty and "Resource Type" in vmmanager_df.columns else pd.DataFrame()

    metrics = [
        ("Projects", int(vm_df["Project"].nunique()) if not vm_df.empty and "Project" in vm_df.columns else 0, "Projects with discovered Compute Engine VMs"),
        ("Compute VMs - all", len(vm_df), "Includes running and stopped instances"),
        ("Compute VMs - running", len(running), "Configuration baseline for currently running instances"),
        ("Compute VMs - stopped/suspended", len(stopped), "Review whether these should migrate"),
        ("vCPU - all VMs", numeric_sum(vm_df, "vCPUs"), "Raw source allocation; not a utilization-based target"),
        ("Memory GiB - all VMs", numeric_sum(vm_df, "Memory GiB"), "Raw source allocation; not a utilization-based target"),
        ("vCPU - running VMs", numeric_sum(running, "vCPUs"), "Raw source allocation for running VMs"),
        ("Memory GiB - running VMs", numeric_sum(running, "Memory GiB"), "Raw source allocation for running VMs"),
        ("Persistent disks", len(disk_df), "Includes attached and unattached persistent disks"),
        ("Persistent disk capacity GiB", numeric_sum(disk_df, "Size GiB"), "Capacity only; collect IOPS/throughput/latency before tier selection"),
        ("External VM IP addresses", int(numeric_sum(vm_df, "External IP Count")), "Review public exposure and Azure ingress/egress design"),
        ("GKE clusters", len(cluster_df), "Candidate AKS clusters; consolidation requires architecture review"),
        ("GKE node pools", len(nodepool_df), "Candidate AKS node pools"),
        ("Discovered GKE nodes", len(gke_nodes), "Mapped through GKE instance groups and VM labels"),
        ("Managed instance groups", len(mig_df), "Candidate VM Scale Sets; GKE-owned groups normally become AKS node pools"),
        ("VMs in managed instance groups", len(mig_vms), "Includes GKE-backed managed groups"),
        ("Cloud SQL instances", len(sql_df), "MySQL, PostgreSQL, and SQL Server database instances"),
        ("Cloud SQL primary instances", len(sql_primaries), "Primary/standalone instances (excludes read replicas)"),
        ("Cloud SQL read replicas", len(sql_replicas), "Map to Azure read replicas where supported"),
        ("Cloud SQL HA instances", len(sql_ha), "Regional (HA) instances; map to Azure zone-redundant HA"),
        ("Cloud SQL storage GiB", numeric_sum(sql_df, "Data Disk Size GiB"), "Provisioned storage; collect growth and IOPS before tier selection"),
        ("Cloud Storage buckets", len(storage_df), "Object storage; collect stored bytes, object count, and egress via Cloud Monitoring"),
        ("BigQuery datasets", len(bigquery_df), "Analytics datasets; confirm query/slot patterns, pipelines, and governance"),
        ("BigQuery logical GiB (sampled)", numeric_sum(bigquery_df, "Total Logical GiB"), "Aggregated table logical size where computed; verify Size Complete column"),
        ("Cloud Logging resources", len(logging_df), "Log routing sinks and log storage buckets"),
        ("Backup and DR resources", len(backupdr_df), "Management servers, backup vaults, and backup plans"),
        ("NetApp volumes", len(netapp_volumes), "Map protocol, service level, and capacity to Azure NetApp Files"),
        ("NetApp volume capacity GiB", numeric_sum(netapp_volumes, "Capacity GiB"), "Provisioned volume capacity; confirm used vs provisioned and snapshots"),
        ("VM Manager patch deployments", len(vm_patch), "Map to Azure Update Manager schedules (Azure Arc for hybrid)"),
        ("Cloud Asset Inventory resources", len(assets_df), "Service count only; service-specific sizing is still required"),
    ]
    return pd.DataFrame(metrics, columns=["Metric", "Value", "Sizing / Migration Note"])


def create_recommendations() -> pd.DataFrame:
    rows = [
        (1, "Performance-based right-sizing", "Collect 14-30 days of CPU utilization, memory working set, disk IOPS/throughput/latency, and network throughput at suitable percentiles. Current recommendations use provisioned configuration only."),
        (2, "GKE workload inventory", "Export namespaces, deployments/statefulsets/daemonsets, pod requests and limits, HPA/VPA, PDBs, ingress, services, storage classes/PVCs, secrets/configmaps, CRDs, and add-ons. Node counts alone cannot size AKS accurately."),
        (3, "Database detail", "Cloud SQL engines, versions, editions, tiers (vCPU/RAM), storage, HA, replicas, backups, flags, and per-instance databases are now collected. Still add Spanner/Bigtable/Firestore/AlloyDB detail plus storage growth, IOPS, RPO/RTO, extensions, connections, and query performance."),
        (4, "Storage consumption", "Cloud Storage bucket configuration (location, class, versioning, lifecycle, retention, encryption, public access) is now collected. Still add stored bytes, object count, request rates, and egress via Cloud Monitoring before selecting Azure tier and redundancy."),
        (5, "Application dependencies", "Use application maps, VPC Flow Logs, load-balancer backends, DNS, service discovery, Kubernetes manifests, and owner interviews to identify communication paths and migration waves."),
        (6, "Network architecture", "Inventory CIDRs, routes, peering, Shared VPC, Cloud NAT, VPN/Interconnect, firewall rules, DNS, load balancers, certificates, bandwidth, latency, and overlapping address space."),
        (7, "Identity and security", "Map IAM roles, service-account usage, workload identity, KMS keys, secrets, org policies, Security Command Center findings, logging, retention, and regulatory controls to Entra ID and Azure Policy."),
        (8, "Availability and resilience", "Capture zones/regions, SLOs, maintenance windows, backup/restore tests, failover, RPO/RTO, autoscaling, quotas, and disaster-recovery topology."),
        (9, "Cost and licensing", "Add current GCP billing export, committed-use discounts, sustained-use effects, Windows/SQL/RHEL/SUSE licensing, support, backup, monitoring, data transfer, and Azure reservations/savings plans."),
        (10, "Azure SKU and quota validation", "Query the intended Azure region for VM/disk/AKS SKU availability, zone support, limits, quotas, and pricing before finalizing the BOQ."),
        (11, "PaaS-specific configuration", "Dedicated collectors now cover Cloud Storage, BigQuery, Cloud Logging, Backup and DR, NetApp Volumes, and VM Manager. Still add service-specific collectors for Cloud Run, Functions, Pub/Sub, Cloud Monitoring, Dataflow, Dataproc, Redis, DNS, KMS, Artifact Registry, and load balancers; asset counts are not enough for sizing."),
        (12, "Regional MIG completeness", "This script queries regional MIGs only in regions where VMs were discovered. Add an all-region scan if zero-instance regional MIGs must also be reported."),
    ]
    return pd.DataFrame(rows, columns=["Priority", "Feature to Add", "Why It Matters"])


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells[:2000]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 80))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="GCP inventory and Azure migration baseline BOQ generator")
    parser.add_argument("--scope", required=True, help="organizations/ID | folders/ID | projects/ID")
    parser.add_argument("--output-prefix", default="gcp_azure_inventory", help="Output path prefix without extension")
    parser.add_argument("--skip-assets", action="store_true", help="Skip Cloud Asset Inventory collection")
    parser.add_argument("--skip-gke", action="store_true", help="Skip GKE cluster/node-pool enrichment")
    parser.add_argument("--skip-sql", action="store_true", help="Skip Cloud SQL database collection")
    parser.add_argument("--skip-storage", action="store_true", help="Skip Cloud Storage bucket collection")
    parser.add_argument("--skip-bigquery", action="store_true", help="Skip BigQuery dataset collection")
    parser.add_argument("--skip-logging", action="store_true", help="Skip Cloud Logging sink/bucket collection")
    parser.add_argument("--skip-backupdr", action="store_true", help="Skip Backup and DR Service collection")
    parser.add_argument("--skip-netapp", action="store_true", help="Skip NetApp Volumes collection")
    parser.add_argument("--skip-vmmanager", action="store_true", help="Skip VM Manager (OS Config) collection")
    args = parser.parse_args()

    errors: List[Dict[str, str]] = []
    credentials, _ = default(scopes=[CLOUD_PLATFORM_SCOPE])
    crm = build("cloudresourcemanager", "v3", credentials=credentials, cache_discovery=False)
    compute = build("compute", "v1", credentials=credentials, cache_discovery=False)
    container = None if args.skip_gke else build("container", "v1", credentials=credentials, cache_discovery=False)
    sqladmin = None if args.skip_sql else build("sqladmin", "v1", credentials=credentials, cache_discovery=False)
    storage = None if args.skip_storage else build("storage", "v1", credentials=credentials, cache_discovery=False)
    bigquery = None if args.skip_bigquery else build("bigquery", "v2", credentials=credentials, cache_discovery=False)
    logging_svc = None if args.skip_logging else build("logging", "v2", credentials=credentials, cache_discovery=False)
    backupdr = None if args.skip_backupdr else build("backupdr", "v1", credentials=credentials, cache_discovery=False)
    netapp = None if args.skip_netapp else build("netapp", "v1", credentials=credentials, cache_discovery=False)
    osconfig = None if args.skip_vmmanager else build("osconfig", "v1", credentials=credentials, cache_discovery=False)

    print("Discovering projects recursively...")
    projects = discover_projects(args.scope, crm, errors)
    project_ids = [p["Project ID"] for p in projects if p.get("Project ID")]
    print(f"Projects discovered: {len(project_ids)}")

    project_number_to_id = {
        str(p.get("Project Number", "")): p.get("Project ID", "")
        for p in projects if p.get("Project Number")
    }

    assets: List[Dict[str, Any]] = []
    if not args.skip_assets:
        print("Collecting Cloud Asset Inventory...")
        try:
            assets = list_assets(args.scope, project_number_to_id)
        except Exception as ex:
            errors.append({"Area": "Cloud Asset Inventory", "Project": args.scope, "Error": str(ex)})
            print(f"Cloud Asset Inventory failed; continuing: {ex}", file=sys.stderr)

    all_vms: List[Dict[str, Any]] = []
    all_disks: List[Dict[str, Any]] = []
    all_nics: List[Dict[str, Any]] = []
    all_migs: List[Dict[str, Any]] = []
    all_clusters: List[Dict[str, Any]] = []
    all_nodepools: List[Dict[str, Any]] = []
    all_sql: List[Dict[str, Any]] = []
    all_storage: List[Dict[str, Any]] = []
    all_bigquery: List[Dict[str, Any]] = []
    all_logging: List[Dict[str, Any]] = []
    all_backupdr: List[Dict[str, Any]] = []
    all_netapp: List[Dict[str, Any]] = []
    all_vmmanager: List[Dict[str, Any]] = []
    machine_cache: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    for index, project in enumerate(project_ids, start=1):
        print(f"[{index}/{len(project_ids)}] Inventorying {project}...")
        disks, disk_lookup = list_disks(project, compute, errors)
        vms, nics, regions = list_instances(project, compute, disk_lookup, machine_cache, errors)
        migs, mig_map = list_managed_instance_groups(project, compute, regions, errors)

        cluster_rows: List[Dict[str, Any]] = []
        nodepool_rows: List[Dict[str, Any]] = []
        gke_map: Dict[Tuple[str, str, str], Dict[str, str]] = {}
        if container is not None:
            cluster_rows, nodepool_rows, gke_map = list_gke(project, container, compute, errors)

        apply_associations(vms, mig_map, gke_map)
        all_disks.extend(disks)
        all_vms.extend(vms)
        all_nics.extend(nics)
        all_migs.extend(migs)
        all_clusters.extend(cluster_rows)
        all_nodepools.extend(nodepool_rows)

        if sqladmin is not None:
            all_sql.extend(list_cloud_sql(project, sqladmin, errors))
        if storage is not None:
            all_storage.extend(list_storage_buckets(project, storage, errors))
        if bigquery is not None:
            all_bigquery.extend(list_bigquery(project, bigquery, errors))
        if logging_svc is not None:
            all_logging.extend(list_cloud_logging(project, logging_svc, errors))
        if backupdr is not None:
            all_backupdr.extend(list_backup_dr(project, backupdr, errors))
        if netapp is not None:
            all_netapp.extend(list_netapp(project, netapp, errors))
        if osconfig is not None:
            all_vmmanager.extend(list_vm_manager(project, osconfig, errors))

    projects_df = dataframe(projects)
    assets_df = dataframe(assets, ["Project ID", "Project Number", "Asset Type", "Name", "Location", "Version", "Ancestors"])
    asset_summary_df = create_asset_summary(assets_df)
    vm_df = dataframe(all_vms)
    disk_df = dataframe(all_disks)
    nic_df = dataframe(all_nics)
    mig_df = dataframe(all_migs)
    cluster_df = dataframe(all_clusters)
    nodepool_df = dataframe(all_nodepools)
    sql_df = dataframe(all_sql)
    storage_df = dataframe(all_storage)
    bigquery_df = dataframe(all_bigquery)
    logging_df = dataframe(all_logging)
    backupdr_df = dataframe(all_backupdr)
    netapp_df = dataframe(all_netapp)
    vmmanager_df = dataframe(all_vmmanager)
    errors_df = dataframe(errors, ["Area", "Project", "Error"])
    recommendations_df = create_recommendations()
    azure_baseline_df = create_azure_baseline(vm_df, disk_df, cluster_df, nodepool_df, mig_df, assets_df, sql_df,
                                              storage_df, bigquery_df, logging_df, backupdr_df, netapp_df, vmmanager_df)

    prefix = Path(args.output_prefix)
    prefix.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path = prefix.with_suffix(".xlsx")

    print("Writing reports...")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        azure_baseline_df.to_excel(writer, sheet_name="Azure_Baseline", index=False)
        asset_summary_df.to_excel(writer, sheet_name="Service_Mapping", index=False)
        projects_df.to_excel(writer, sheet_name="Projects", index=False)
        assets_df.to_excel(writer, sheet_name="Assets", index=False)
        vm_df.to_excel(writer, sheet_name="Compute_VMs", index=False)
        disk_df.to_excel(writer, sheet_name="Disks", index=False)
        nic_df.to_excel(writer, sheet_name="NICs", index=False)
        mig_df.to_excel(writer, sheet_name="Managed_Groups", index=False)
        cluster_df.to_excel(writer, sheet_name="GKE_Clusters", index=False)
        nodepool_df.to_excel(writer, sheet_name="GKE_NodePools", index=False)
        sql_df.to_excel(writer, sheet_name="Cloud_SQL", index=False)
        storage_df.to_excel(writer, sheet_name="Cloud_Storage", index=False)
        bigquery_df.to_excel(writer, sheet_name="BigQuery", index=False)
        logging_df.to_excel(writer, sheet_name="Cloud_Logging", index=False)
        backupdr_df.to_excel(writer, sheet_name="Backup_DR", index=False)
        netapp_df.to_excel(writer, sheet_name="NetApp_Volumes", index=False)
        vmmanager_df.to_excel(writer, sheet_name="VM_Manager", index=False)
        recommendations_df.to_excel(writer, sheet_name="Recommendations", index=False)
        errors_df.to_excel(writer, sheet_name="Errors", index=False)

    format_workbook(xlsx_path)

    csv_outputs = {
        f"{prefix}_assets.csv": assets_df,
        f"{prefix}_asset_summary.csv": asset_summary_df,
        f"{prefix}_compute.csv": vm_df,
        f"{prefix}_disks.csv": disk_df,
        f"{prefix}_gke_nodepools.csv": nodepool_df,
        f"{prefix}_cloud_sql.csv": sql_df,
        f"{prefix}_cloud_storage.csv": storage_df,
        f"{prefix}_bigquery.csv": bigquery_df,
        f"{prefix}_cloud_logging.csv": logging_df,
        f"{prefix}_backup_dr.csv": backupdr_df,
        f"{prefix}_netapp_volumes.csv": netapp_df,
        f"{prefix}_vm_manager.csv": vmmanager_df,
    }
    for filename, df in csv_outputs.items():
        df.to_csv(filename, index=False)

    print("\n=== Azure Migration Baseline ===")
    if not azure_baseline_df.empty:
        print(azure_baseline_df.to_string(index=False))

    print("\nGenerated:")
    print(f" - {xlsx_path}")
    for filename in csv_outputs:
        print(f" - {filename}")
    if errors:
        print(f"\nCompleted with {len(errors)} warning/error entries. Review the Errors sheet.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
